# -*- coding: utf-8 -*-
"""
Genera "SPORT FACTORY_KPI 2024.xlsx" con el mismo formato que los archivos de
2025/2026, tomando los KPIs y el margen directamente de la base de datos ICG
[NEWSPF] (SQL Server local).

KPIs por tienda y mes:  VENTAS, UDS, FACT, UPT, VPT, MARG% (margen bruto %).
  VENTAS = SUM(ALBVENTALIN.TOTAL)              (neto de devoluciones y descuentos)
  UDS    = SUM(ALBVENTALIN.UNIDADESTOTAL)
  FACT   = COUNT(DISTINCT ticket)              (NUMSERIE-NUMALBARAN-N)
  UPT    = UDS / FACT
  VPT    = VENTAS / FACT
  MARG%  = (VENTAS - SUM(COSTE*UNIDADESTOTAL)) / VENTAS * 100

Metodología validada contra el Excel manual (A2 ene-2026: 95,493.75 vs 95,493.43).
"""
import os
import pyodbc
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

YEAR = 2024
HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "SPORT FACTORY_KPI %d.xlsx" % YEAR)

CONN_STR = (
    "DRIVER={ODBC Driver 17 for SQL Server};"
    "SERVER=localhost;DATABASE=NEWSPF;"
    "Trusted_Connection=yes;TrustServerCertificate=yes;"
)

MONTHS = ["ENE", "FEB", "MAR", "ABR", "MAY", "JUN",
          "JUL", "AGO", "SEP", "OCT", "NOV", "DIC"]
KPI_HEADERS = ["VENTAS", "UDS", "FACT", "UPT ", "VPT", "MARG%"]

# (etiqueta en el Excel, CODALMACEN en ICG)
NB_STORES = [("NB MP", "A2"), ("NB ALB", "A3"), ("NB DOR", "A4"), ("NB MM", "A7")]
RB_STORES = [("RB MP", "B1"), ("RB ALB", "B2")]


def fetch_raw():
    """{(codalmacen, mes): dict(ventas,uds,fact,costo)}"""
    sql = """
        SELECT lin.CODALMACEN AS alm, MONTH(cab.FECHA) AS mes,
               SUM(lin.TOTAL) AS ventas,
               SUM(lin.UNIDADESTOTAL) AS uds,
               COUNT(DISTINCT CONCAT(cab.NUMSERIE COLLATE DATABASE_DEFAULT,'-',
                     cab.NUMALBARAN,'-',cab.N COLLATE DATABASE_DEFAULT)) AS fact,
               SUM(lin.COSTE*lin.UNIDADESTOTAL) AS costo
        FROM ALBVENTACAB cab
        JOIN ALBVENTALIN lin ON lin.NUMSERIE=cab.NUMSERIE
             AND lin.NUMALBARAN=cab.NUMALBARAN AND lin.N=cab.N
        WHERE YEAR(cab.FECHA)=? AND lin.CODALMACEN IN ('A2','A3','A4','A7','B1','B2')
        GROUP BY lin.CODALMACEN, MONTH(cab.FECHA)
    """
    conn = pyodbc.connect(CONN_STR)
    cur = conn.cursor()
    cur.execute(sql, [YEAR])
    raw = {}
    for alm, mes, ventas, uds, fact, costo in cur.fetchall():
        raw[(alm, mes)] = {
            "ventas": float(ventas or 0), "uds": float(uds or 0),
            "fact": int(fact or 0), "costo": float(costo or 0),
        }
    conn.close()
    return raw


def kpis_for(raw, code, mes):
    r = raw.get((code, mes))
    if not r or r["ventas"] == 0:
        return None
    v, u, f, c = r["ventas"], r["uds"], r["fact"], r["costo"]
    return {
        "ventas": v, "uds": u, "fact": f,
        "upt": (u / f) if f else None,
        "vpt": (v / f) if f else None,
        "marg": ((v - c) / v * 100) if v else None,
    }


def totals_for(raw, codes, mes):
    v = u = f = c = 0.0
    got = False
    for code in codes:
        r = raw.get((code, mes))
        if r and r["ventas"]:
            v += r["ventas"]; u += r["uds"]; f += r["fact"]; c += r["costo"]; got = True
    if not got:
        return None
    return {"ventas": v, "uds": u, "fact": int(f),
            "upt": (u / f) if f else None, "vpt": (v / f) if f else None,
            "marg": ((v - c) / v * 100) if v else None}


def row_vals(get_month):
    """Build the flat list of cells for one store row across all months + ACUM."""
    out = []
    acc = {"ventas": 0.0, "uds": 0.0, "fact": 0.0, "costo_num": 0.0}
    # For ACUM margin we need a weighted figure: accumulate ventas & margin$.
    acc_marg_dollars = 0.0
    for mi in range(1, 13):
        k = get_month(mi)
        if k:
            out += [k["ventas"], k["uds"], k["fact"], k["upt"], k["vpt"], k["marg"]]
            acc["ventas"] += k["ventas"]; acc["uds"] += k["uds"]; acc["fact"] += k["fact"]
            acc_marg_dollars += k["ventas"] * (k["marg"] or 0) / 100.0
        else:
            out += [None, None, None, None, None, None]
    # ACUM block
    v, u, f = acc["ventas"], acc["uds"], acc["fact"]
    if v:
        out += [v, u, int(f), (u / f) if f else None, (v / f) if f else None,
                (acc_marg_dollars / v * 100) if v else None]
    else:
        out += [None] * 6
    return out


def main():
    raw = fetch_raw()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = str(YEAR)

    STRIDE = len(KPI_HEADERS)  # 6
    # Title
    ws.cell(1, 1, "SPORT FACTORY ").font = Font(bold=True, size=14)

    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font = Font(bold=True, color="FFFFFF")
    mon_font = Font(bold=True)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def month_cols():
        # column (1-based) where each month block starts; +1 because col1 is label
        return [2 + i * STRIDE for i in range(12)] + [2 + 12 * STRIDE]  # last = ACUM

    def write_headers(month_row, kpi_row):
        starts = month_cols()
        labels = MONTHS + ["ACUM"]
        for start, lab in zip(starts, labels):
            c = ws.cell(month_row, start, lab)
            c.font = mon_font; c.alignment = Alignment(horizontal="center")
            for j, kh in enumerate(KPI_HEADERS):
                cc = ws.cell(kpi_row, start + j, kh)
                cc.font = hdr_font; cc.fill = hdr_fill
                cc.alignment = Alignment(horizontal="center")

    def write_row(r, label, vals, bold=False):
        ws.cell(r, 1, label).font = Font(bold=bold or label == "TOTAL")
        for j, v in enumerate(vals):
            cell = ws.cell(r, 2 + j, v)
            # number format by KPI position within stride
            pos = j % STRIDE
            if pos in (0, 4):   # ventas, vpt -> money
                cell.number_format = '#,##0.00'
            elif pos in (1, 2):  # uds, fact -> int
                cell.number_format = '#,##0'
            elif pos == 3:       # upt -> 2 dec
                cell.number_format = '0.000'
            elif pos == 5:       # marg%
                cell.number_format = '0.0"%"'

    # ---- NB block ----
    write_headers(3, 4)
    ws.cell(4, 1, "TNB").font = hdr_font
    ws.cell(4, 1).fill = hdr_fill
    r = 5
    for label, code in NB_STORES:
        write_row(r, label, row_vals(lambda mi, c=code: kpis_for(raw, c, mi)))
        r += 1
    write_row(r, "TOTAL", row_vals(lambda mi: totals_for(raw, [c for _, c in NB_STORES], mi)), bold=True)
    r += 2

    # ---- RB block ----
    kpi_row = r; month_row = r - 1
    write_headers(month_row, kpi_row)
    ws.cell(kpi_row, 1, "TRB").font = hdr_font
    ws.cell(kpi_row, 1).fill = hdr_fill
    r = kpi_row + 1
    for label, code in RB_STORES:
        write_row(r, label, row_vals(lambda mi, c=code: kpis_for(raw, c, mi)))
        r += 1
    write_row(r, "TOTAL", row_vals(lambda mi: totals_for(raw, [c for _, c in RB_STORES], mi)), bold=True)

    ws.freeze_panes = "B5"
    ws.column_dimensions["A"].width = 10
    wb.save(OUT)
    print("Escrito:", OUT)

    # quick summary
    for label, code in NB_STORES + RB_STORES:
        got = [MONTHS[mi-1] for mi in range(1, 13) if kpis_for(raw, code, mi)]
        print("  %-7s (%s): %s" % (label, code, ", ".join(got) or "-"))


if __name__ == "__main__":
    main()
