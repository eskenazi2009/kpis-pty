# -*- coding: utf-8 -*-
"""
Build a standalone, offline, password-protected KPI dashboard (dashboard.html)
for the Sport Factory Panama stores.

Sources:
  - "SPORT FACTORY_KPI <year>.xlsx"  -> VENTAS/UDS/FACT/UPT/VPT per store/month.
  - ICG [NEWSPF] (SQL Server)        -> MARG% (gross margin %) per store/month.
  - "Presupuesto de ventas SF 2026_JP_actualizado.xlsx" (sheet
    "Presupuesto de vts 2026") -> monthly budget (the blue "2026" column) per
    NB store, used for the "Presupuesto vs Real" tab.

The data + budget are encrypted (AES-256-GCM, key from a passphrase via
PBKDF2-SHA256) and embedded; the page decrypts in the browser with Web Crypto.
Charts are hand-rolled inline SVG (no CDN, fully offline).

Re-run whenever the spreadsheets change (it will ask for the passphrase):
    python build_dashboard.py
"""

import json
import glob
import os
import re
import base64
import getpass
import secrets
import openpyxl
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.ciphers.aead import AESGCM

HERE = os.path.dirname(os.path.abspath(__file__))
PBKDF2_ITERATIONS = 200000  # must match the value used in the browser

MONTHS = ["ENE", "FEB", "MAR", "ABR", "MAY", "JUN",
          "JUL", "AGO", "SEP", "OCT", "NOV", "DIC"]
MONTH_SET = set(MONTHS)
BRAND_HEADERS = {"TNB": "NB", "TRB": "RB"}

# Store label (as shown in the KPI xlsx) -> ICG CODALMACEN
STORE_CODE = {"NB MP": "A2", "NB ALB": "A3", "NB DOR": "A4", "NB MM": "A7",
              "RB MP": "B1", "RB ALB": "B2"}
NB_CODES = ["A2", "A3", "A4", "A7"]
RB_CODES = ["B1", "B2"]

CONN_STR = (
    "DRIVER={ODBC Driver 17 for SQL Server};"
    "SERVER=localhost;DATABASE=NEWSPF;"
    "Trusted_Connection=yes;TrustServerCertificate=yes;"
)

BUDGET_FILE = "Presupuesto de ventas SF 2026_JP_actualizado.xlsx"
BUDGET_SHEET = "Presupuesto de vts 2026"
BUDGET_YEAR = "2026"
STORE_LABELS = set(STORE_CODE.keys())  # NB/RB store labels used as budget headers


def encrypt_payload(plaintext, passphrase):
    """Encrypt with AES-256-GCM using a PBKDF2-SHA256 derived key.
    Layout is byte-for-byte compatible with the browser Web Crypto API."""
    salt = secrets.token_bytes(16)
    kdf = PBKDF2HMAC(algorithm=hashes.SHA256(), length=32,
                     salt=salt, iterations=PBKDF2_ITERATIONS)
    key = kdf.derive(passphrase.encode("utf-8"))
    iv = secrets.token_bytes(12)
    ct = AESGCM(key).encrypt(iv, plaintext.encode("utf-8"), None)  # ciphertext||tag
    b64 = lambda b: base64.b64encode(b).decode("ascii")
    return {"salt": b64(salt), "iv": b64(iv), "ct": b64(ct), "iter": PBKDF2_ITERATIONS}


def num(v):
    """Return a float for a real numeric cell, else None."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "").strip())
    except (ValueError, TypeError):
        return None


def parse_workbook(path):
    """Parse one KPI workbook -> { store: { MONTH: {ventas,uds,fact,upt,vpt} } }."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # 1) Find month start columns from the first row that holds month names.
    month_cols = {}
    for r in range(1, min(ws.max_row, 6) + 1):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(r, c).value
            if isinstance(val, str) and val.strip().upper() in MONTH_SET:
                month_cols[val.strip().upper()] = c
        if month_cols:
            break
    month_cols = {m: c for m, c in month_cols.items() if m in MONTH_SET}

    # 2) Walk rows; track current brand from TNB/TRB header rows.
    stores = {}
    brand = None
    for r in range(1, ws.max_row + 1):
        raw = ws.cell(r, 1).value
        label = (str(raw).strip() if raw is not None else "")
        if not label:
            continue
        up = label.upper()
        if up in BRAND_HEADERS:
            brand = BRAND_HEADERS[up]
            continue
        if up in ("SPORT FACTORY", "TNB", "TRB"):
            continue
        store = (brand or "?") + " TOTAL" if up == "TOTAL" else label

        months = {}
        for m, base in month_cols.items():
            ventas = num(ws.cell(r, base).value)
            if ventas is None or ventas == 0:
                continue
            months[m] = {
                "ventas": ventas,
                "uds":    num(ws.cell(r, base + 1).value),
                "fact":   num(ws.cell(r, base + 2).value),
                "upt":    num(ws.cell(r, base + 3).value),
                "vpt":    num(ws.cell(r, base + 4).value),
            }
        if months:
            stores[store] = months
    return stores


def year_from_name(path):
    m = re.search(r"(20\d{2})", os.path.basename(path))
    return m.group(1) if m else os.path.splitext(os.path.basename(path))[0]


def build_data():
    files = sorted(glob.glob(os.path.join(HERE, "SPORT FACTORY_KPI *.xlsx")))
    data = {}
    for f in files:
        if os.path.basename(f).startswith("~$"):
            continue
        data[year_from_name(f)] = parse_workbook(f)
    return data


def fetch_margins(years):
    """Gross margin % per year/store/month from ICG, incl. NB/RB TOTAL.
    Returns {year: {store: {MONTH: pct}}}. Empty dict if the DB is unreachable."""
    try:
        import pyodbc
        conn = pyodbc.connect(CONN_STR, timeout=5)
    except Exception as e:
        print("WARN: no se pudo conectar a la BD; el margen se omite (%s)" % e)
        return {}

    cur = conn.cursor()
    codes = list(STORE_CODE.values())
    ph = ",".join("?" * len(codes))
    rev = {c: lbl for lbl, c in STORE_CODE.items()}
    result = {}
    for y in years:
        if not str(y).isdigit():
            continue
        cur.execute(
            "SELECT lin.CODALMACEN, MONTH(cab.FECHA), SUM(lin.TOTAL), "
            "SUM(lin.COSTE*lin.UNIDADESTOTAL) "
            "FROM ALBVENTACAB cab JOIN ALBVENTALIN lin "
            "  ON lin.NUMSERIE=cab.NUMSERIE AND lin.NUMALBARAN=cab.NUMALBARAN AND lin.N=cab.N "
            "WHERE YEAR(cab.FECHA)=? AND lin.CODALMACEN IN (%s) "
            "GROUP BY lin.CODALMACEN, MONTH(cab.FECHA)" % ph,
            [int(y)] + codes)
        raw = {}
        for alm, mes, v, c in cur.fetchall():
            raw[(alm, int(mes))] = (float(v or 0), float(c or 0))
        ym = {}
        for (alm, mes), (v, c) in raw.items():
            lbl = rev.get(alm)
            if lbl and v:
                ym.setdefault(lbl, {})[MONTHS[mes - 1]] = round((v - c) / v * 100, 2)
        for tot, tcodes in (("NB TOTAL", NB_CODES), ("RB TOTAL", RB_CODES)):
            for mi in range(1, 13):
                v = sum(raw.get((a, mi), (0, 0))[0] for a in tcodes)
                c = sum(raw.get((a, mi), (0, 0))[1] for a in tcodes)
                if v:
                    ym.setdefault(tot, {})[MONTHS[mi - 1]] = round((v - c) / v * 100, 2)
        result[str(y)] = ym
    conn.close()
    return result


def parse_budget():
    """Monthly budget (blue '2026' column) per store -> {year, stores}.

    The sheet stacks two tables: NB stores (MP/ALB/DOR/MM + TOTAL) on top and
    RB stores (MP/ALB + TOTAL) below. Each store header spans [2025, 2026, REAL],
    so the budget column is (header column + 1). We detect header rows by the
    store labels they contain, then read the month rows that follow."""
    path = os.path.join(HERE, BUDGET_FILE)
    empty = {"year": None, "stores": {}}
    if not os.path.exists(path):
        print("WARN: no se encontró el archivo de presupuesto; pestaña vacía.")
        return empty
    wb = openpyxl.load_workbook(path, data_only=True)
    if BUDGET_SHEET not in wb.sheetnames:
        print("WARN: no existe la hoja '%s'; pestaña vacía." % BUDGET_SHEET)
        return empty
    ws = wb[BUDGET_SHEET]
    mon3 = {m[:3]: m for m in MONTHS}
    stores = {}
    colmap = {}   # {budget_column: store_label}, rebuilt at each header row
    for r in range(1, ws.max_row + 1):
        rowvals = {c: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
        labels = [(c, v.strip()) for c, v in rowvals.items()
                  if isinstance(v, str) and v.strip() in STORE_LABELS]
        if labels:                         # a store-header row -> reset colmap
            brand = "NB" if any(l.startswith("NB") for _, l in labels) else "RB"
            colmap = {c + 1: lbl for c, lbl in labels}
            for c, v in rowvals.items():   # the "TOTAL" header in the same block
                if isinstance(v, str) and v.strip().upper() == "TOTAL":
                    colmap[c + 1] = brand + " TOTAL"
            continue
        mv = rowvals.get(1)
        if isinstance(mv, str) and colmap:
            mon = mon3.get(mv.strip()[:3].upper())
            if mon:
                for col, lbl in colmap.items():
                    v = num(ws.cell(r, col).value)
                    if v is not None and v != 0:
                        stores.setdefault(lbl, {})[mon] = v
    stores = {k: v for k, v in stores.items() if v}
    return {"year": BUDGET_YEAR, "stores": stores}


def merge_margins(data, margins):
    for y, stores in data.items():
        ym = margins.get(y, {})
        for store, months in stores.items():
            sm = ym.get(store, {})
            for mon, rec in months.items():
                if mon in sm:
                    rec["margen"] = sm[mon]


HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Sport Factory Panamá — KPI Dashboard</title>
<style>
  :root{
    --bg:#0f1419; --panel:#1a2230; --panel2:#222c3d; --line:#2e3a4d;
    --text:#e8edf4; --muted:#93a1b5; --accent:#4da3ff; --accent2:#ff8c42;
    --good:#36d399; --bad:#f87272; --budget:#7c8aa0;
  }
  *{box-sizing:border-box}
  body{margin:0;background:var(--bg);color:var(--text);
       font-family:Segoe UI,Roboto,Helvetica,Arial,sans-serif;font-size:14px}
  header{padding:18px 24px;border-bottom:1px solid var(--line);
         display:flex;align-items:baseline;gap:14px;flex-wrap:wrap}
  header h1{font-size:20px;margin:0;font-weight:600}
  header .sub{color:var(--muted);font-size:13px}
  .wrap{padding:18px 24px;max-width:1280px;margin:0 auto}
  .tabs{display:flex;gap:4px;margin-bottom:18px;border-bottom:1px solid var(--line)}
  .tabbtn{background:none;border:none;color:var(--muted);padding:10px 18px;font-size:14px;
          cursor:pointer;border-bottom:2px solid transparent;font-weight:600}
  .tabbtn.active{color:var(--text);border-bottom-color:var(--accent)}
  .tabbtn:hover{color:var(--text)}
  .controls{display:flex;gap:18px;flex-wrap:wrap;align-items:flex-end;margin-bottom:16px}
  .ctl{display:flex;flex-direction:column;gap:5px}
  .ctl label{font-size:12px;color:var(--muted);text-transform:uppercase;letter-spacing:.04em}
  select{background:var(--panel2);color:var(--text);border:1px solid var(--line);
         border-radius:8px;padding:8px 12px;font-size:14px;min-width:170px}
  .stores-wrap{margin-bottom:20px}
  .stores-wrap .lab{font-size:12px;color:var(--muted);text-transform:uppercase;
                    letter-spacing:.04em;margin-bottom:7px}
  .stores{display:flex;flex-wrap:wrap;gap:8px}
  .sb{background:var(--panel2);border:1px solid var(--line);color:var(--text);
      border-radius:8px;padding:8px 14px;cursor:pointer;font-size:13px;transition:.12s}
  .sb:hover{border-color:var(--accent)}
  .sb.active{background:var(--accent);border-color:var(--accent);color:#04263f;font-weight:700}
  .sb.total{border-style:dashed}
  .cards{display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:12px;margin-bottom:22px}
  .card{background:var(--panel);border:1px solid var(--line);border-radius:12px;padding:14px}
  .card .k{font-size:12px;color:var(--muted);text-transform:uppercase;letter-spacing:.04em}
  .card .v{font-size:22px;font-weight:700;margin-top:6px}
  .card .prev{font-size:12px;color:var(--muted);margin-top:2px}
  .delta{font-size:13px;font-weight:600;margin-top:6px}
  .delta.up{color:var(--good)} .delta.down{color:var(--bad)} .delta.flat{color:var(--muted)}
  .card.sel{outline:2px solid var(--accent)}
  .card.click{cursor:pointer}
  .panel{background:var(--panel);border:1px solid var(--line);border-radius:12px;
         padding:16px 18px;margin-bottom:22px}
  .panel h2{font-size:15px;margin:0 0 14px;font-weight:600}
  .panel h2 .hint{color:var(--muted);font-weight:400;font-size:12px;margin-left:8px}
  .chart-grid{display:grid;grid-template-columns:1fr 1fr;gap:22px}
  @media(max-width:980px){.chart-grid{grid-template-columns:1fr}}
  .chartbox{position:relative}
  svg{width:100%;height:auto;display:block}
  .hz{cursor:crosshair}
  .legend{display:flex;gap:18px;margin-top:8px;font-size:12px;color:var(--muted);flex-wrap:wrap}
  .legend i{display:inline-block;width:11px;height:11px;border-radius:3px;margin-right:6px;vertical-align:-1px}
  table{width:100%;border-collapse:collapse;font-size:13px}
  th,td{padding:9px 12px;text-align:right;border-bottom:1px solid var(--line)}
  th:first-child,td:first-child{text-align:left}
  th{color:var(--muted);font-weight:600;cursor:pointer;user-select:none;text-transform:uppercase;
     font-size:11px;letter-spacing:.04em;white-space:nowrap}
  th.sortdir::after{content:" \25BE";opacity:.7}
  th.sortdir.asc::after{content:" \25B4"}
  tbody tr:hover{background:var(--panel2)}
  td.pos{color:var(--good)} td.neg{color:var(--bad)}
  .foot{color:var(--muted);font-size:12px;margin-top:8px}
  .tag{display:inline-block;background:var(--panel2);border:1px solid var(--line);
       border-radius:999px;padding:2px 10px;font-size:12px;color:var(--muted)}
  #tip{position:fixed;display:none;pointer-events:none;z-index:50;background:#0b0f15;
       border:1px solid var(--line);border-radius:8px;padding:8px 10px;font-size:12px;
       color:var(--text);box-shadow:0 6px 20px rgba(0,0,0,.45);min-width:130px}
  #tip .tt{color:var(--muted);font-size:11px;margin-bottom:5px}
  #tip .row{display:flex;justify-content:space-between;gap:16px;line-height:1.55}
  #tip .row i{display:inline-block;width:9px;height:9px;border-radius:2px;margin-right:6px}
  #lock{position:fixed;inset:0;display:flex;align-items:center;justify-content:center;
        background:var(--bg);z-index:100;padding:20px}
  .lockbox{background:var(--panel);border:1px solid var(--line);border-radius:16px;
           padding:32px 28px;max-width:380px;width:100%;text-align:center}
  .lockbox .ico{font-size:34px;line-height:1}
  .lockbox h2{font-size:17px;margin:14px 0 4px;font-weight:600}
  .lockbox p{color:var(--muted);font-size:13px;margin:0 0 18px}
  .lockbox input{width:100%;background:var(--panel2);border:1px solid var(--line);color:var(--text);
                 border-radius:9px;padding:11px 13px;font-size:15px;margin-bottom:12px}
  .lockbox input:focus{outline:none;border-color:var(--accent)}
  .lockbox button{width:100%;background:var(--accent);color:#04263f;border:none;border-radius:9px;
                  padding:11px;font-size:15px;font-weight:700;cursor:pointer}
  .lockbox button:disabled{opacity:.6;cursor:default}
  .lockerr{color:var(--bad);font-size:13px;margin-top:12px;display:none}
</style>
</head>
<body>
<header>
  <h1>Sport Factory Panamá</h1>
  <span class="sub">KPI Dashboard</span>
  <span class="tag" id="yearsTag"></span>
</header>
<div id="lock">
  <div class="lockbox">
    <div class="ico">🔒</div>
    <h2>Sport Factory Panamá — KPI Dashboard</h2>
    <p>Ingresa la contraseña para ver el tablero.</p>
    <input id="pw" type="password" placeholder="Contraseña" autocomplete="off" autocapitalize="off" spellcheck="false">
    <button id="unlockBtn">Entrar</button>
    <div id="lockErr" class="lockerr"></div>
  </div>
</div>

<div class="wrap" id="app" style="display:none">
  <div class="tabs">
    <button class="tabbtn active" data-tab="kpis">KPIs</button>
    <button class="tabbtn" data-tab="budget" id="tabBudget">Presupuesto vs Real</button>
  </div>

  <div class="controls">
    <div class="ctl" id="ctlKpi">
      <label for="kpi">Indicador (KPI)</label>
      <select id="kpi"></select>
    </div>
    <div class="ctl" id="ctlYears">
      <label>Años (compara 1–3)</label>
      <div class="stores" id="years"></div>
    </div>
    <div class="ctl">
      <label>Período</label>
      <div class="stores" id="period"></div>
    </div>
  </div>

  <div class="stores-wrap">
    <div class="lab">Tienda</div>
    <div class="stores" id="stores"></div>
  </div>

  <!-- ============ KPIs tab ============ -->
  <div id="tab-kpis">
    <div class="cards" id="cards"></div>
    <div class="chart-grid">
      <div class="panel">
        <h2>Tendencia mensual <span class="hint" id="trendHint"></span></h2>
        <div class="chartbox" id="trend"></div>
        <div class="legend" id="trendLegend"></div>
      </div>
      <div class="panel">
        <h2>Comparación año contra año <span class="hint" id="yoyHint"></span></h2>
        <div class="chartbox" id="yoy"></div>
        <div class="legend" id="yoyLegend"></div>
      </div>
    </div>
    <div class="panel">
      <h2>Comparación por tienda <span class="hint" id="tableHint"></span></h2>
      <table><thead><tr id="thr"></tr></thead><tbody id="tb"></tbody></table>
      <div class="foot" id="tblFoot"></div>
    </div>
  </div>

  <!-- ============ Presupuesto vs Real tab ============ -->
  <div id="tab-budget" style="display:none">
    <div class="cards" id="bcards"></div>
    <div class="panel">
      <h2>Presupuesto vs Real por mes <span class="hint" id="budHint"></span></h2>
      <div class="chartbox" id="budchart"></div>
      <div class="legend" id="budLegend"></div>
    </div>
    <div class="panel">
      <h2>Cumplimiento por tienda <span class="hint" id="budTableHint"></span></h2>
      <table><thead><tr id="bthr"></tr></thead><tbody id="btb"></tbody></table>
      <div class="foot" id="budFoot"></div>
    </div>
  </div>

  <div class="foot">Generado desde los archivos .xlsx y la base de datos ICG · funciona sin conexión.</div>
</div>
<div id="tip"></div>

<script>
const ENCRYPTED = /*__ENC__*/;
let DATA = {};
let BUDGET = {year:null, stores:{}};
let YEARS = [];
const MONTHS = /*__MONTHS__*/;
const MONTH_LABEL = {ENE:"Ene",FEB:"Feb",MAR:"Mar",ABR:"Abr",MAY:"May",JUN:"Jun",
                     JUL:"Jul",AGO:"Ago",SEP:"Sep",OCT:"Oct",NOV:"Nov",DIC:"Dic"};
const SEM = {all:MONTHS, s1:["ENE","FEB","MAR","ABR","MAY","JUN"], s2:["JUL","AGO","SEP","OCT","NOV","DIC"]};
const KPIS = [
  {id:"ventas", label:"VENTAS", desc:"Ingreso total ($)",     fmt:"money", color:"#4da3ff"},
  {id:"uds",    label:"UDS",    desc:"Unidades vendidas",     fmt:"int",   color:"#ff8c42"},
  {id:"fact",   label:"FACT",   desc:"Cantidad de facturas",  fmt:"int",   color:"#36d399"},
  {id:"upt",    label:"UPT",    desc:"Unidades por factura",  fmt:"dec2",  color:"#c084fc"},
  {id:"vpt",    label:"VPT",    desc:"Valor ($) por factura", fmt:"money", color:"#fbbf24"},
  {id:"margen", label:"MARGEN %", desc:"Margen bruto (%)",    fmt:"pct1",  color:"#22d3ee"},
];
const KMAP = Object.fromEntries(KPIS.map(k=>[k.id,k]));
const YEAR_COLORS = ["#4da3ff","#ff8c42","#36d399","#c084fc","#fbbf24"]; // by year index
function yearColor(y){ return YEAR_COLORS[Math.max(0,YEARS.indexOf(y))%YEAR_COLORS.length]; }
const BCOL = {p:"#7c8aa0", r:"#4da3ff"};      // presupuesto, real

// ---- formatting -----------------------------------------------------------
function fmtVal(v, kind){
  if(v===null||v===undefined||isNaN(v)) return "—";
  if(kind==="money") return "$"+v.toLocaleString("en-US",{maximumFractionDigits:0});
  if(kind==="int")   return Math.round(v).toLocaleString("en-US");
  if(kind==="dec2")  return v.toFixed(2);
  if(kind==="pct1")  return v.toFixed(1)+"%";
  return String(v);
}
function pct(a,b){
  if(a===null||b===null||a===0||a===undefined||b===undefined) return null;
  return (b-a)/a*100;
}
function deltaHtml(p){
  if(p===null) return '<div class="delta flat">—</div>';
  const cls = p>0.05?"up":(p<-0.05?"down":"flat");
  const arrow = p>0.05?"▲":(p<-0.05?"▼":"▬");
  return `<div class="delta ${cls}">${arrow} ${p>=0?"+":""}${p.toFixed(1)}%</div>`;
}

// ---- period / aggregation -------------------------------------------------
function periodMonths(){ return SEM[state.period] || MONTHS; }
function periodLabel(){ return state.period==="s1"?"Q1 · Ene–Jun":
                               state.period==="s2"?"Q2 · Jul–Dic":"Año completo"; }
function monthsWithData(year, store){
  const s = (DATA[year]||{})[store]||{};
  return MONTHS.filter(m=>s[m]);
}
function selectedYears(){ return YEARS.filter(y=>state.years.includes(y)); }
// months (within the period) present in EVERY selected year -> like-for-like
function comparableMonths(store){
  const ys=selectedYears();
  return periodMonths().filter(m=> ys.every(y=> (DATA[y]||{})[store] && (DATA[y]||{})[store][m]));
}
function ytd(year, store, kpi, months){
  const s=(DATA[year]||{})[store]||{};
  let ventas=0,uds=0,fact=0,margd=0,hasM=false,n=0;
  for(const m of months){ if(!s[m]) continue; n++;
    ventas+=s[m].ventas||0; uds+=s[m].uds||0; fact+=s[m].fact||0;
    if(s[m].margen!=null){ margd+=(s[m].ventas||0)*s[m].margen; hasM=true; } }
  if(!n) return null;
  switch(kpi){
    case "ventas": return ventas;
    case "uds":    return uds;
    case "fact":   return fact;
    case "upt":    return fact? uds/fact : null;
    case "vpt":    return fact? ventas/fact : null;
    case "margen": return (hasM && ventas)? margd/ventas : null;
  }
}

// ---- state ----------------------------------------------------------------
const state = {tab:"kpis", period:"all", kpi:"ventas", store:null, years:[],
               sortKey:null, sortDir:-1};

function allStores(){
  const set=[];
  for(const y of YEARS) for(const s of Object.keys(DATA[y])) if(!set.includes(s)) set.push(s);
  return set;
}
function budgetStores(){
  const order=["NB MP","NB ALB","NB DOR","NB MM","NB TOTAL","RB MP","RB ALB","RB TOTAL"];
  const have=BUDGET.stores||{};
  return order.filter(s=>have[s] && Object.keys(have[s]).length);
}

// ---- tooltip / hover ------------------------------------------------------
function tipHtml(title, rows){
  return `<div class="tt">${title}</div>`+rows.map(r=>
    `<div class="row"><span>${r.color?`<i style="background:${r.color}"></i>`:""}${r.label}</span><b>${r.val}</b></div>`).join("");
}
function showTip(html, ev){
  const t=document.getElementById("tip");
  t.innerHTML=html; t.style.display="block";
  const pad=14; let x=ev.clientX+pad, y=ev.clientY+pad;
  const r=t.getBoundingClientRect();
  if(x+r.width>window.innerWidth)  x=ev.clientX-r.width-pad;
  if(y+r.height>window.innerHeight) y=ev.clientY-r.height-pad;
  t.style.left=x+"px"; t.style.top=y+"px";
}
function hideTip(){ document.getElementById("tip").style.display="none"; }
function attachHover(container, model){
  const svg=container.querySelector("svg");
  const ns="http://www.w3.org/2000/svg";
  const guide=document.createElementNS(ns,"line");
  guide.setAttribute("stroke","#93a1b5"); guide.setAttribute("stroke-width","1");
  guide.setAttribute("stroke-dasharray","3 3"); guide.style.display="none";
  guide.setAttribute("y1",model.guideTop); guide.setAttribute("y2",model.guideBottom);
  svg.appendChild(guide);
  container.querySelectorAll(".hz").forEach(z=>{
    const i=+z.dataset.i;
    z.addEventListener("mouseenter",()=>{ guide.style.display="";
      guide.setAttribute("x1",model.xs[i]); guide.setAttribute("x2",model.xs[i]); });
    z.addEventListener("mousemove",e=>showTip(model.months[i].html,e));
    z.addEventListener("mouseleave",()=>{ guide.style.display="none"; hideTip(); });
  });
}

// ---- SVG helpers ----------------------------------------------------------
function svgEl(w,h){ return {w,h,parts:[],
  add(s){this.parts.push(s);},
  out(){return `<svg viewBox="0 0 ${this.w} ${this.h}" preserveAspectRatio="xMidYMid meet">${this.parts.join("")}</svg>`;}};}
function txt(x,y,s,opt={}){const a=opt.anchor||"middle";const fs=opt.size||11;
  const col=opt.color||"#93a1b5";const w=opt.weight||"400";
  return `<text x="${x}" y="${y}" text-anchor="${a}" font-size="${fs}" fill="${col}" font-weight="${w}">${s}</text>`;}
function niceMax(v){
  if(v<=0) return 1;
  const pow=Math.pow(10,Math.floor(Math.log10(v)));
  const n=v/pow; const step = n<=1?1:n<=2?2:n<=5?5:10;
  return step*pow;
}

// ---- KPI cards ------------------------------------------------------------
function renderCards(){
  const wrap=document.getElementById("cards"); wrap.innerHTML="";
  const ys=selectedYears();
  const last=ys[ys.length-1], prev=ys.length>=2?ys[ys.length-2]:null;
  const cm = comparableMonths(state.store);
  const periodLbl = cm.length ? `${MONTH_LABEL[cm[0]]}–${MONTH_LABEL[cm[cm.length-1]]}` : "—";
  KPIS.forEach(k=>{
    const bv=ytd(last, state.store, k.id, cm);
    // delta of the newest selected year vs the previous selected year
    let deltaMarkup="";
    if(prev){
      const pv=ytd(prev, state.store, k.id, cm);
      if(k.id==="margen"){
        const d=(pv!=null&&bv!=null)?(bv-pv):null;
        deltaMarkup = d===null?'<div class="delta flat">—</div>':
          `<div class="delta ${d>0.05?"up":d<-0.05?"down":"flat"}">${d>=0?"+":""}${d.toFixed(1)} pp <span style="color:var(--muted);font-weight:400">vs ${prev}</span></div>`;
      } else {
        const p=pct(pv,bv);
        deltaMarkup = p===null?'<div class="delta flat">—</div>':
          `<div class="delta ${p>0.05?"up":p<-0.05?"down":"flat"}">${p>=0?"+":""}${p.toFixed(1)}% <span style="color:var(--muted);font-weight:400">vs ${prev}</span></div>`;
      }
    }
    const others=ys.slice(0,-1).map(y=>`<div class="prev">${y}: ${fmtVal(ytd(y,state.store,k.id,cm),k.fmt)}</div>`).join("");
    const div=document.createElement("div");
    div.className="card click"+(k.id===state.kpi?" sel":"");
    div.innerHTML=`<div class="k">${k.label}</div>
      <div class="v">${fmtVal(bv,k.fmt)}</div>
      <div class="prev">${last} · ${periodLbl}</div>
      ${deltaMarkup}${others}`;
    div.onclick=()=>{state.kpi=k.id; render();};
    wrap.appendChild(div);
  });
}

// ---- trend (line, single KPI, both years) ---------------------------------
function renderTrend(){
  const k=KMAP[state.kpi], XM=periodMonths(), ys=selectedYears();
  document.getElementById("trendHint").textContent = `${k.label} · ${state.store} · ${periodLabel()}`;
  const W=560,H=300,PL=64,PR=16,PT=16,PB=34;
  const series = ys.map(y=>({year:y, s:(DATA[y]||{})[state.store]||{}, color:yearColor(y)}));
  const vals=[];
  series.forEach(se=> XM.forEach(m=>{ if(se.s[m]&&se.s[m][k.id]!=null) vals.push(se.s[m][k.id]); }));
  const max = niceMax(vals.length?Math.max(...vals):1);
  const plotW=W-PL-PR, step=plotW/Math.max(1,(XM.length-1));
  const x=i=> PL + i*step, y=v=> H-PB - (v/max)*(H-PT-PB);
  const g=svgEl(W,H);
  for(let t=0;t<=4;t++){const yy=PT+(H-PT-PB)*t/4; const val=max*(1-t/4);
    g.add(`<line x1="${PL}" y1="${yy}" x2="${W-PR}" y2="${yy}" stroke="#2e3a4d" stroke-width="1"/>`);
    g.add(txt(PL-8,yy+4,fmtVal(val,k.fmt),{anchor:"end",size:10}));}
  XM.forEach((m,i)=>g.add(txt(x(i),H-PB+18,MONTH_LABEL[m],{size:10})));
  series.forEach(se=>{
    let d="",pts="",started=false;
    XM.forEach((m,i)=>{const rec=se.s[m]; if(rec&&rec[k.id]!=null){
      const X=x(i),Y=y(rec[k.id]); d+=(started?"L":"M")+X+" "+Y+" "; started=true;
      pts+=`<circle cx="${X}" cy="${Y}" r="3.2" fill="${se.color}"/>`;}});
    if(d) g.add(`<path d="${d}" fill="none" stroke="${se.color}" stroke-width="2.4"/>`);
    g.add(pts);
  });
  const model={xs:XM.map((m,i)=>x(i)), guideTop:PT, guideBottom:H-PB, months:[]};
  XM.forEach((m,i)=>{
    const rows=series.filter(se=>se.s[m]&&se.s[m][k.id]!=null)
      .map(se=>({color:se.color,label:se.year,val:fmtVal(se.s[m][k.id],k.fmt)}));
    if(rows.length){
      model.months[i]={html:tipHtml(`${MONTH_LABEL[m]} · ${k.label}`,rows)};
      const zx=Math.max(PL,x(i)-step/2), zw=Math.min(step,W-PR-zx);
      g.add(`<rect class="hz" data-i="${i}" x="${zx}" y="${PT}" width="${zw}" height="${H-PT-PB}" fill="#000" fill-opacity="0" pointer-events="all"/>`);
    } else model.months[i]=null;
  });
  const c=document.getElementById("trend"); c.innerHTML=g.out(); attachHover(c,model);
  document.getElementById("trendLegend").innerHTML=
    series.map(se=>`<span><i style="background:${se.color}"></i>${se.year}</span>`).join("");
}

// ---- YoY (grouped bars) ---------------------------------------------------
function renderYoY(){
  const k=KMAP[state.kpi], ys=selectedYears();
  document.getElementById("yoyHint").textContent = `${k.label} · ${state.store} · ${periodLabel()}`;
  const cm = comparableMonths(state.store);
  const W=560,H=300,PL=64,PR=16,PT=16,PB=34;
  const g=svgEl(W,H);
  if(!cm.length){ g.add(txt(W/2,H/2,"Sin meses comparables",{size:13}));
    document.getElementById("yoy").innerHTML=g.out();
    document.getElementById("yoyLegend").innerHTML=""; return; }
  const S = ys.map(y=>({year:y, s:DATA[y][state.store], color:yearColor(y)}));
  const vals=[]; cm.forEach(m=> S.forEach(se=>vals.push(se.s[m][k.id])));
  const max=niceMax(Math.max(...vals));
  const y=v=> H-PB - (v/max)*(H-PT-PB);
  const band=(W-PL-PR)/cm.length;
  const total=S.length;
  const barW=Math.max(6,Math.min(28,(band*0.8)/total));
  const space=Math.min(4,band*0.04);
  const groupW=total*barW+(total-1)*space;
  for(let t=0;t<=4;t++){const yy=PT+(H-PT-PB)*t/4; const val=max*(1-t/4);
    g.add(`<line x1="${PL}" y1="${yy}" x2="${W-PR}" y2="${yy}" stroke="#2e3a4d" stroke-width="1"/>`);
    g.add(txt(PL-8,yy+4,fmtVal(val,k.fmt),{anchor:"end",size:10}));}
  const model={xs:[], guideTop:PT, guideBottom:H-PB, months:[]};
  cm.forEach((m,i)=>{
    const cx=PL+band*i+band/2, x0=cx-groupW/2;
    S.forEach((se,j)=>{ const v=se.s[m][k.id], bx=x0+j*(barW+space);
      g.add(`<rect x="${bx}" y="${y(v)}" width="${barW}" height="${H-PB-y(v)}" fill="${se.color}" rx="2"/>`); });
    g.add(txt(cx,H-PB+18,MONTH_LABEL[m],{size:10}));
    model.xs.push(cx);
    const rows=S.map(se=>({color:se.color,label:se.year,val:fmtVal(se.s[m][k.id],k.fmt)}));
    if(S.length>=2){
      const a=S[0].s[m][k.id], b=S[S.length-1].s[m][k.id];
      const d = k.id==="margen" ? (b-a) : pct(a,b);
      rows.push({color:"",label:`Δ ${S[0].year}→${S[S.length-1].year}`,
        val:(d===null?"—":(d>=0?"+":"")+d.toFixed(1)+(k.id==="margen"?" pp":"%"))});
    }
    model.months.push({html:tipHtml(`${MONTH_LABEL[m]} · ${k.label}`,rows)});
    g.add(`<rect class="hz" data-i="${i}" x="${PL+band*i}" y="${PT}" width="${band}" height="${H-PT-PB}" fill="#000" fill-opacity="0" pointer-events="all"/>`);
  });
  const c=document.getElementById("yoy"); c.innerHTML=g.out(); attachHover(c,model);
  document.getElementById("yoyLegend").innerHTML=
    S.map(se=>`<span><i style="background:${se.color}"></i>${se.year}</span>`).join("");
}

// ---- store comparison table ----------------------------------------------
function renderTable(){
  const k=KMAP[state.kpi], ys=selectedYears(), isMarg=k.id==="margen";
  const showDelta = ys.length>=2;
  const yKey=y=>"y_"+y;
  const cols=[{key:"store",label:"Tienda"}].concat(ys.map(y=>({key:yKey(y),label:y})));
  if(showDelta) cols.push({key:"delta",label:isMarg?"Δ pp":"Δ%"});
  // keep sort key valid for the current columns
  if(!cols.some(c=>c.key===state.sortKey)) state.sortKey = yKey(ys[ys.length-1]);

  const thr=document.getElementById("thr"); thr.innerHTML="";
  cols.forEach(c=>{
    const th=document.createElement("th"); th.textContent=c.label; th.dataset.key=c.key;
    if(c.key===state.sortKey){th.classList.add("sortdir"); if(state.sortDir>0)th.classList.add("asc");}
    th.onclick=()=>{ if(state.sortKey===c.key) state.sortDir*=-1;
      else {state.sortKey=c.key; state.sortDir=(c.key==="store")?1:-1;} render(); };
    thr.appendChild(th);
  });
  const rows=allStores().map(s=>{
    const cm=comparableMonths(s);
    const r={store:s};
    ys.forEach(y=> r[yKey(y)]=ytd(y,s,k.id,cm));
    if(showDelta){ const a=r[yKey(ys[0])], b=r[yKey(ys[ys.length-1])];
      r.delta = isMarg ? ((a!=null&&b!=null)?(b-a):null) : pct(a,b); }
    return r;
  });
  const sk=state.sortKey, dir=state.sortDir;
  rows.sort((a,b)=>{ let x=a[sk], y=b[sk];
    if(sk==="store") return dir*String(x).localeCompare(String(y));
    if(x===null||x===undefined) return 1; if(y===null||y===undefined) return -1; return dir*(x-y); });
  const tb=document.getElementById("tb"); tb.innerHTML="";
  rows.forEach(r=>{
    const tr=document.createElement("tr");
    let cells=`<td>${r.store}${r.store===state.store?' &#9679;':''}</td>`;
    ys.forEach(y=> cells+=`<td>${fmtVal(r[yKey(y)],k.fmt)}</td>`);
    if(showDelta){ const dcls=r.delta===null?"":(r.delta>0?"pos":(r.delta<0?"neg":""));
      const dtxt=r.delta===null?"—":(r.delta>=0?"+":"")+r.delta.toFixed(1)+(isMarg?" pp":"%");
      cells+=`<td class="${dcls}">${dtxt}</td>`; }
    tr.innerHTML=cells; tr.style.cursor="pointer";
    tr.onclick=()=>{state.store=r.store; render();};
    tb.appendChild(tr);
  });
  const cm=comparableMonths(state.store);
  const per = cm.length?`${MONTH_LABEL[cm[0]]}–${MONTH_LABEL[cm[cm.length-1]]}`:"—";
  document.getElementById("tableHint").textContent=`${k.label} · acumulado comparable (${per})`;
  document.getElementById("tblFoot").textContent = isMarg
    ? "El margen acumulado se pondera por ventas (no es promedio simple); la variación (pp) compara el año más antiguo vs el más reciente seleccionado."
    : "Los acumulados (YTD) comparan solo los meses presentes en todos los años seleccionados. UPT/VPT se recalculan sobre los acumulados.";
}

// ---- Presupuesto vs Real --------------------------------------------------
function budgetData(store){
  const bs=(BUDGET.stores||{})[store]||{};
  const real=(DATA[BUDGET.year]||{})[store]||{};
  const pm=periodMonths();
  const withBudget = pm.filter(m=>bs[m]!=null);          // budget exists (chart)
  const elapsed    = withBudget.filter(m=>real[m]);      // real also exists (KPIs)
  return {bs, real, withBudget, elapsed};
}
function renderBudget(){
  const store=state.store, by=BUDGET.year;
  const {bs, real, withBudget, elapsed} = budgetData(store);
  const presup = elapsed.reduce((t,m)=>t+bs[m],0);
  const realv  = elapsed.reduce((t,m)=>t+(real[m]?real[m].ventas:0),0);
  const cumpl  = presup? realv/presup*100 : null;
  const dif    = realv - presup;
  const perLbl = elapsed.length?`${MONTH_LABEL[elapsed[0]]}–${MONTH_LABEL[elapsed[elapsed.length-1]]}`:"—";

  // cards
  const cumplCls = cumpl===null?"flat":(cumpl>=100?"up":(cumpl>=90?"flat":"down"));
  document.getElementById("bcards").innerHTML = `
    <div class="card"><div class="k">Real</div><div class="v">${fmtVal(realv||null,"money")}</div>
      <div class="prev">${by} · ${perLbl}</div></div>
    <div class="card"><div class="k">Presupuesto</div><div class="v">${fmtVal(presup||null,"money")}</div>
      <div class="prev">mismos meses</div></div>
    <div class="card"><div class="k">Cumplimiento</div><div class="v">${cumpl===null?"—":cumpl.toFixed(1)+"%"}</div>
      <div class="delta ${cumplCls}">${cumpl===null?"":(cumpl>=100?"▲ meta alcanzada":"▼ bajo meta")}</div></div>
    <div class="card"><div class="k">Diferencia</div><div class="v">${fmtVal(dif,"money")}</div>
      <div class="delta ${dif>=0?"up":"down"}">${dif>=0?"+":""}${fmtVal(Math.abs(dif),"money").replace("$","$")}</div></div>`;

  // chart: grouped bars presupuesto vs real over months that have a budget
  document.getElementById("budHint").textContent = `${store} · ${by} · ${periodLabel()}`;
  const W=1140,H=320,PL=64,PR=16,PT=16,PB=34;
  const g=svgEl(W,H);
  if(!withBudget.length){
    g.add(txt(W/2,H/2,"Sin presupuesto para esta selección",{size:13}));
    document.getElementById("budchart").innerHTML=g.out();
    document.getElementById("budLegend").innerHTML="";
  } else {
    const vals=[]; withBudget.forEach(m=>{ vals.push(bs[m]); if(real[m]) vals.push(real[m].ventas); });
    const max=niceMax(Math.max(...vals));
    const y=v=> H-PB - (v/max)*(H-PT-PB);
    const band=(W-PL-PR)/withBudget.length, bw=Math.min(30,band/3);
    for(let t=0;t<=4;t++){const yy=PT+(H-PT-PB)*t/4; const val=max*(1-t/4);
      g.add(`<line x1="${PL}" y1="${yy}" x2="${W-PR}" y2="${yy}" stroke="#2e3a4d" stroke-width="1"/>`);
      g.add(txt(PL-8,yy+4,fmtVal(val,"money"),{anchor:"end",size:10}));}
    const model={xs:[], guideTop:PT, guideBottom:H-PB, months:[]};
    withBudget.forEach((m,i)=>{
      const cx=PL+band*i+band/2, p=bs[m], rv=real[m]?real[m].ventas:null;
      g.add(`<rect x="${cx-bw-2}" y="${y(p)}" width="${bw}" height="${H-PB-y(p)}" fill="${BCOL.p}" rx="2"/>`);
      if(rv!=null) g.add(`<rect x="${cx+2}" y="${y(rv)}" width="${bw}" height="${H-PB-y(rv)}" fill="${BCOL.r}" rx="2"/>`);
      g.add(txt(cx,H-PB+18,MONTH_LABEL[m],{size:10}));
      model.xs.push(cx);
      const cp = (rv!=null)? rv/p*100 : null;
      model.months.push({html:tipHtml(`${MONTH_LABEL[m]} · ${by}`,[
        {color:BCOL.p,label:"Presupuesto",val:fmtVal(p,"money")},
        {color:BCOL.r,label:"Real",val:fmtVal(rv,"money")},
        {color:"",label:"Cumpl.",val:(cp===null?"—":cp.toFixed(1)+"%")},
      ])});
      g.add(`<rect class="hz" data-i="${i}" x="${PL+band*i}" y="${PT}" width="${band}" height="${H-PT-PB}" fill="#000" fill-opacity="0" pointer-events="all"/>`);
    });
    const c=document.getElementById("budchart"); c.innerHTML=g.out(); attachHover(c,model);
    document.getElementById("budLegend").innerHTML=
      `<span><i style="background:${BCOL.p}"></i>Presupuesto</span>
       <span><i style="background:${BCOL.r}"></i>Real</span>`;
  }

  // table: per store cumplimiento over elapsed months
  const thr=document.getElementById("bthr");
  thr.innerHTML=`<th>Tienda</th><th>Presupuesto</th><th>Real</th><th>Cumpl.%</th><th>Dif $</th>`;
  const tb=document.getElementById("btb"); tb.innerHTML="";
  budgetStores().forEach(s=>{
    const d=budgetData(s);
    const p=d.elapsed.reduce((t,m)=>t+d.bs[m],0);
    const rv=d.elapsed.reduce((t,m)=>t+(d.real[m]?d.real[m].ventas:0),0);
    const cp=p?rv/p*100:null, df=rv-p;
    const tr=document.createElement("tr"); tr.style.cursor="pointer";
    tr.innerHTML=`<td>${s}${s===store?' &#9679;':''}</td>
      <td>${fmtVal(p||null,"money")}</td><td>${fmtVal(rv||null,"money")}</td>
      <td class="${cp===null?"":cp>=100?"pos":"neg"}">${cp===null?"—":cp.toFixed(1)+"%"}</td>
      <td class="${df>=0?"pos":"neg"}">${df>=0?"+":""}${fmtVal(Math.abs(df),"money")}</td>`;
    tr.onclick=()=>{state.store=s; render();};
    tb.appendChild(tr);
  });
  document.getElementById("budTableHint").textContent = `${periodLabel()} · ${perLbl}`;
  document.getElementById("budFoot").textContent =
    "Presupuesto = columna mensual del archivo de presupuesto. Real = ventas reales. El cumplimiento compara solo los meses con datos reales.";
}

// ---- store buttons / period / tabs ---------------------------------------
function renderStoreButtons(){
  const wrap=document.getElementById("stores"); wrap.innerHTML="";
  const list = state.tab==="budget" ? budgetStores() : allStores();
  list.forEach(s=>{
    const b=document.createElement("button");
    b.className="sb"+(s===state.store?" active":"")+(s.includes("TOTAL")?" total":"");
    b.textContent=s;
    b.onclick=()=>{state.store=s; render();};
    wrap.appendChild(b);
  });
}
function renderPeriod(){
  const wrap=document.getElementById("period"); wrap.innerHTML="";
  [["all","Año completo"],["s1","Q1 · Ene–Jun"],["s2","Q2 · Jul–Dic"]].forEach(([v,lbl])=>{
    const b=document.createElement("button");
    b.className="sb"+(state.period===v?" active":""); b.textContent=lbl;
    b.onclick=()=>{state.period=v; render();};
    wrap.appendChild(b);
  });
}
function toggleYear(y){
  if(state.years.includes(y)){
    if(state.years.length>1) state.years=state.years.filter(x=>x!==y);  // keep at least 1
  } else {
    state.years=YEARS.filter(x=>state.years.includes(x)||x===y);        // keep chronological
  }
  render();
}
function renderYearButtons(){
  const wrap=document.getElementById("years"); wrap.innerHTML="";
  YEARS.forEach(y=>{
    const b=document.createElement("button");
    b.className="sb"+(state.years.includes(y)?" active":""); b.textContent=y;
    b.onclick=()=>toggleYear(y);
    wrap.appendChild(b);
  });
}
function switchTab(t){
  state.tab=t;
  if(t==="budget"){ const bs=budgetStores();
    if(!bs.includes(state.store)) state.store = bs.includes("NB TOTAL")?"NB TOTAL":bs[0]; }
  render();
}

function render(){
  document.querySelectorAll(".tabbtn").forEach(b=>b.classList.toggle("active",b.dataset.tab===state.tab));
  const kpisTab = state.tab==="kpis";
  document.getElementById("ctlKpi").style.display = kpisTab?"":"none";
  document.getElementById("ctlYears").style.display = kpisTab?"":"none";
  document.getElementById("tab-kpis").style.display = kpisTab?"":"none";
  document.getElementById("tab-budget").style.display = kpisTab?"none":"";
  renderPeriod();
  renderYearButtons();
  renderStoreButtons();
  if(kpisTab){ renderCards(); renderTrend(); renderYoY(); renderTable(); }
  else { renderBudget(); }
}

// ---- controls -------------------------------------------------------------
function initControls(){
  const kpiSel=document.getElementById("kpi");
  KPIS.forEach(k=>kpiSel.add(new Option(`${k.label} — ${k.desc}`, k.id)));
  kpiSel.value=state.kpi;
  kpiSel.onchange=e=>{state.kpi=e.target.value; render();};

  document.querySelectorAll(".tabbtn").forEach(b=> b.onclick=()=>switchTab(b.dataset.tab));
  if(!budgetStores().length) document.getElementById("tabBudget").style.display="none";

  state.store = allStores()[0];
  document.getElementById("yearsTag").textContent = YEARS.join(" · ");
}

// ---- decryption / unlock --------------------------------------------------
async function decryptData(passphrase){
  const dec = s => Uint8Array.from(atob(s), c => c.charCodeAt(0));
  const salt = dec(ENCRYPTED.salt), iv = dec(ENCRYPTED.iv), ct = dec(ENCRYPTED.ct);
  const baseKey = await crypto.subtle.importKey(
    "raw", new TextEncoder().encode(passphrase), "PBKDF2", false, ["deriveKey"]);
  const key = await crypto.subtle.deriveKey(
    {name:"PBKDF2", salt, iterations:ENCRYPTED.iter, hash:"SHA-256"},
    baseKey, {name:"AES-GCM", length:256}, false, ["decrypt"]);
  const plain = await crypto.subtle.decrypt({name:"AES-GCM", iv}, key, ct);
  return JSON.parse(new TextDecoder().decode(plain));
}
function boot(payload){
  DATA = payload.data || payload;               // backward compatible
  BUDGET = payload.budget || {year:null, stores:{}};
  YEARS = Object.keys(DATA).sort();
  state.years = YEARS.slice(-2);                 // default: two most recent years
  initControls();
  render();
  document.getElementById("lock").style.display = "none";
  document.getElementById("app").style.display = "";
}
async function unlock(){
  const inp = document.getElementById("pw");
  const err = document.getElementById("lockErr");
  const btn = document.getElementById("unlockBtn");
  if(!window.crypto || !crypto.subtle){
    err.textContent = "Este navegador no permite descifrar desde un archivo local. Abre el enlace web (https).";
    err.style.display = "block"; return;
  }
  btn.disabled = true; err.style.display = "none";
  try{ boot(await decryptData(inp.value)); }
  catch(e){ err.textContent = "Contraseña incorrecta."; err.style.display = "block"; inp.select(); }
  finally{ btn.disabled = false; }
}
document.getElementById("unlockBtn").onclick = unlock;
document.getElementById("pw").addEventListener("keydown", e=>{ if(e.key==="Enter") unlock(); });
document.getElementById("pw").focus();
</script>
</body>
</html>
"""


def main():
    data = build_data()
    if not data:
        raise SystemExit("No 'SPORT FACTORY_KPI *.xlsx' files found in this folder.")

    margins = fetch_margins(list(data.keys()))
    merge_margins(data, margins)
    budget = parse_budget()

    passphrase = os.environ.get("DASH_PASSPHRASE")
    if not passphrase:
        passphrase = getpass.getpass("Contraseña para el tablero: ")
    if not passphrase:
        raise SystemExit("A passphrase is required to encrypt the dashboard.")

    payload = {"data": data, "budget": budget}
    enc = encrypt_payload(json.dumps(payload, ensure_ascii=False), passphrase)
    html = (HTML_TEMPLATE
            .replace("/*__ENC__*/", json.dumps(enc))
            .replace("/*__MONTHS__*/", json.dumps(MONTHS)))
    for name in ("dashboard.html", "index.html"):
        with open(os.path.join(HERE, name), "w", encoding="utf-8") as f:
            f.write(html)

    print("Wrote dashboard.html and index.html")
    for y in sorted(data):
        stores = data[y]
        months = sorted({m for s in stores.values() for m in s}, key=MONTHS.index)
        nmarg = sum(1 for s in stores.values() for r in s.values() if "margen" in r)
        print(f"  {y}: {len(stores)} filas, meses: {', '.join(months) or '—'} | celdas con margen: {nmarg}")
    print(f"  Presupuesto: {len(budget.get('stores',{}))} tiendas ({budget.get('year')})")


if __name__ == "__main__":
    main()
